"""
Processeur de fichiers Excel et CSV
Lecture et extraction de contenu pour .xlsx, .xls, .csv
"""

import csv
import os
from pathlib import Path
from typing import Any, Dict, List


class ExcelProcessor:
    """Processeur pour les fichiers Excel (.xlsx, .xls) et CSV (.csv)"""

    def __init__(self):
        self.supported_extensions = [".xlsx", ".xls", ".csv"]
        self._check_dependencies()

    def _check_dependencies(self):
        try:
            import openpyxl  # noqa: F401
            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False
            print("⚠️ Module 'openpyxl' non installé. pip install openpyxl")

        try:
            import xlrd  # noqa: F401
            self.xlrd_available = True
        except ImportError:
            self.xlrd_available = False

    # ------------------------------------------------------------------ #
    #  Méthode principale                                                   #
    # ------------------------------------------------------------------ #

    def read_excel(self, file_path: str) -> Dict[str, Any]:
        """Lit un fichier Excel ou CSV et retourne le contenu structuré."""
        if not os.path.exists(file_path):
            return {"success": False, "error": f"Fichier non trouvé : {file_path}"}

        ext = Path(file_path).suffix.lower()

        if ext == ".csv":
            return self._read_csv(file_path)
        elif ext == ".xlsx":
            return self._read_xlsx(file_path)
        elif ext == ".xls":
            return self._read_xls(file_path)
        else:
            return {"success": False, "error": f"Format non supporté : {ext}"}

    def extract_text(self, file_path: str) -> Dict[str, Any]:
        """Interface compatible avec les autres processeurs : retourne {success, content}."""
        result = self.read_excel(file_path)
        if result.get("success"):
            return {"success": True, "content": result["content"]}
        return result

    # ------------------------------------------------------------------ #
    #  Lecteurs par format                                                  #
    # ------------------------------------------------------------------ #

    def _read_csv(self, file_path: str) -> Dict[str, Any]:
        """Lit un fichier CSV (encodage automatique)."""
        rows: List[List[str]] = []
        encoding = "utf-8-sig"
        try:
            with open(file_path, newline="", encoding=encoding, errors="replace") as f:
                reader = csv.reader(f)
                for row in reader:
                    rows.append(row)
        except Exception as exc:
            return {"success": False, "error": f"Erreur lecture CSV : {exc}"}

        text = self._rows_to_text("Données CSV", rows)
        return {
            "success": True,
            "content": text,
            "sheets": {"CSV": rows},
            "sheet_names": ["CSV"],
            "total_rows": len(rows),
            "processor": "csv-stdlib",
        }

    def _read_xlsx(self, file_path: str) -> Dict[str, Any]:
        """Lit un fichier .xlsx via openpyxl."""
        if not self.openpyxl_available:
            return {
                "success": False,
                "error": "openpyxl requis pour les fichiers .xlsx. Installez avec : pip install openpyxl",
            }
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True)
            texts: List[str] = []
            sheets_data: Dict[str, List] = {}
            total_rows = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(
                        [str(cell) if cell is not None else "" for cell in row]
                    )
                sheets_data[sheet_name] = rows
                total_rows += len(rows)
                texts.append(self._rows_to_text(sheet_name, rows))

            return {
                "success": True,
                "content": "\n\n".join(texts),
                "sheets": sheets_data,
                "sheet_names": list(wb.sheetnames),
                "total_rows": total_rows,
                "processor": "openpyxl",
            }
        except Exception as exc:
            return {"success": False, "error": f"Erreur lecture .xlsx : {exc}"}

    def _read_xls(self, file_path: str) -> Dict[str, Any]:
        """Lit un fichier .xls (ancien format) via xlrd, sinon tente openpyxl."""
        if self.xlrd_available:
            try:
                import xlrd

                wb = xlrd.open_workbook(file_path)
                texts: List[str] = []
                sheets_data: Dict[str, List] = {}
                total_rows = 0

                for sheet_name in wb.sheet_names():
                    ws = wb.sheet_by_name(sheet_name)
                    rows = []
                    for rx in range(ws.nrows):
                        rows.append([str(ws.cell_value(rx, cx)) for cx in range(ws.ncols)])
                    sheets_data[sheet_name] = rows
                    total_rows += len(rows)
                    texts.append(self._rows_to_text(sheet_name, rows))

                return {
                    "success": True,
                    "content": "\n\n".join(texts),
                    "sheets": sheets_data,
                    "sheet_names": wb.sheet_names(),
                    "total_rows": total_rows,
                    "processor": "xlrd",
                }
            except Exception as exc:
                return {"success": False, "error": f"Erreur lecture .xls (xlrd) : {exc}"}

        # Fallback : openpyxl peut lire certains .xls
        if self.openpyxl_available:
            return self._read_xlsx(file_path)

        return {
            "success": False,
            "error": "Aucune bibliothèque disponible pour .xls. Installez xlrd ou openpyxl.",
        }

    # ------------------------------------------------------------------ #
    #  Formatage texte                                                      #
    # ------------------------------------------------------------------ #

    def _rows_to_text(self, sheet_name: str, rows: List[List[str]]) -> str:
        """Convertit une liste de lignes en texte lisible (format table)."""
        if not rows:
            return f"=== Feuille : {sheet_name} ===\n(vide)"

        # Calculer la largeur max de chaque colonne (limité à 30 chars)
        ncols = max(len(r) for r in rows) if rows else 0
        col_widths = [0] * ncols
        for row in rows:
            for ci, cell in enumerate(row):
                col_widths[ci] = min(30, max(col_widths[ci], len(str(cell))))

        def fmt_row(row: List[str]) -> str:
            cells = []
            for ci in range(ncols):
                val = str(row[ci]) if ci < len(row) else ""
                if len(val) > 30:
                    val = val[:27] + "..."
                cells.append(val.ljust(col_widths[ci]))
            return "| " + " | ".join(cells) + " |"

        separator = "|-" + "-|-".join("-" * w for w in col_widths) + "-|"

        lines = [f"=== Feuille : {sheet_name} ==="]
        max_display = 200  # limite pour éviter de saturer le contexte

        for i, row in enumerate(rows[:max_display]):
            lines.append(fmt_row(row))
            if i == 0 and len(rows) > 1:
                lines.append(separator)

        if len(rows) > max_display:
            lines.append(
                f"... ({len(rows) - max_display} lignes supplémentaires non affichées)"
            )

        lines.append(
            f"\n📊 Total : {len(rows)} lignes × {ncols} colonnes"
        )
        return "\n".join(lines)

    def is_supported(self, file_path: str) -> bool:
        return Path(file_path).suffix.lower() in self.supported_extensions
